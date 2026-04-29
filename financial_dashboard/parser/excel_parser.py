"""Core Excel parsing logic for MSE financial statements.

Parses .xlsx and .xls files downloaded from members.mse.mn containing
Mongolian-language financial statements and extracts structured data.

Real MSE file layout (confirmed from members.mse.mn):
  - Column A (0): sometimes company info row, otherwise empty
  - Column B (1): row numbers like "1.1.1"
  - Column C (2): Mongolian header text (Үзүүлэлт)
  - Column D (3): Previous year values (Эхний үлдэгдэл)
  - Column E (4): Current year values (Эцсийн үлдэгдэл)

Uses openpyxl for .xlsx files and xlrd for .xls files.
"""

import io
import logging
import re
from datetime import datetime

import xlrd
from openpyxl import load_workbook

from .header_mappings import (
    HEADERS_BY_TYPE,
    SHEET_TYPE_PATTERNS,
    match_header,
    normalize_header,
)

# Lazy import to avoid circular deps — only used for sector-aware header selection
def _lookup_sector_key(company_name: str, filename: str = "") -> str:
    """Return a sector key for _SECTOR_SHEET_OVERRIDES lookup.

    For Finance companies with a sub_sector, returns "Finance:{sub_sector}"
    (e.g. "Finance:Securities").  Otherwise returns the plain sector string.
    """
    try:
        from ..scraper.registry_loader import (
            find_sector as _fs,
            find_sub_sector as _fss,
            find_sector_from_filename as _fsf,
        )
        sector = _fs(company_name) or (filename and _fsf(filename)) or ""
        if sector == "Finance":
            sub = _fss(company_name)
            if sub:
                return f"Finance:{sub}"
        return sector
    except Exception:
        return ""


# Back-compat alias used by tests / external callers
def _lookup_sector(company_name: str, filename: str = "") -> str:
    key = _lookup_sector_key(company_name, filename)
    return key.split(":")[0]  # strip sub-sector


# Sector key → which sheet types to use when standard sheet names are detected.
# Keys may be plain sector ("Banking") or composite ("Finance:Securities").
_SECTOR_SHEET_OVERRIDES: dict[str, dict[str, str]] = {
    "Banking": {
        "balance_sheet":   "bank_balance_sheet",
        "income_statement": "bank_income_statement",
    },
    "Insurance": {
        "balance_sheet":   "insurance_balance_sheet",
        "income_statement": "insurance_income_statement",
    },
    # Lending NBFIs (ББСБ, MIK mortgage, etc.) use BoM reporting format —
    # same terminology as banks — so bank header dictionaries apply.
    "Finance": {
        "balance_sheet":   "bank_balance_sheet",
        "income_statement": "bank_income_statement",
    },
    # Securities brokers and investment companies file under FRC format which
    # uses standard Mongolian BS/IS terminology plus securities-specific items.
    "Finance:Securities": {
        "balance_sheet":   "securities_balance_sheet",
        "income_statement": "securities_income_statement",
    },
}

# Column indices in MSE financial statement files
HEADER_COL = 2   # Column C: Mongolian header text
PREV_COL = 3     # Column D: Previous year value
CURR_COL = 4     # Column E: Current year value


def parse_excel_file(file_bytes: bytes, filename: str, sector: str = "") -> dict:
    """Parse an Excel file containing MSE financial statements.

    Args:
        file_bytes: Raw bytes of the uploaded Excel file.
        filename: Original filename (used to extract company/year metadata).
        sector: Optional sector hint ("Banking", "Insurance", or ""). When supplied,
            standard sheet names (balance_sheet / income_statement) are redirected to
            the sector-specific header dictionaries so bank/insurance-specific line
            items are captured.

    Returns:
        Structured dict with metadata and parsed financial data.

    Raises:
        ValueError: If the file cannot be parsed or contains no recognizable data.
    """
    is_xls = filename.lower().endswith(".xls")

    if is_xls:
        return _parse_xls(file_bytes, filename, sector)
    else:
        return _parse_xlsx(file_bytes, filename, sector)


def _parse_xlsx(file_bytes: bytes, filename: str, sector: str = "") -> dict:
    """Parse a .xlsx file using openpyxl."""
    try:
        wb = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as e:
        raise ValueError(f"Cannot open Excel file: {e}")

    company, year = _extract_metadata_from_filename(filename)

    # Detect reporting unit from first sheet before building result dict
    unit_multiplier = 1_000
    if wb.sheetnames:
        unit_multiplier = _detect_reporting_unit_openpyxl(wb[wb.sheetnames[0]])

    result = _make_result_dict(filename, company, year, unit_multiplier)

    # Pre-scan sheets for company name so sector lookup works even when the
    # filename is an opaque MSE code (e.g. "56320254report.xls").
    if not sector:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            extracted = _extract_company_from_openpyxl_sheet(ws)
            if extracted:
                company = extracted
                result["metadata"]["company"] = extracted
                break

    resolved_sector_key = _lookup_sector_key(company, filename) if not sector else sector
    sector_overrides = _SECTOR_SHEET_OVERRIDES.get(resolved_sector_key, {})

    try:
        for sheet_name in wb.sheetnames:
            sheet_type = _detect_sheet_type(sheet_name)
            if not sheet_type:
                continue

            # Redirect standard sheet types to sector-specific ones when sector is known.
            # E.g., a bank's "баланс" sheet is parsed with BANK_BALANCE_SHEET_HEADERS so
            # bank-specific terms like "зээлийн багц" (loan portfolio) are captured.
            effective_type = sector_overrides.get(sheet_type, sheet_type)

            ws = wb[sheet_name]
            # 7 separate dictionaries (one per statement type) prevent false matches:
            # bank terminology overlaps with standard terminology in Mongolian. Merged into
            # one dict, "зээл" (loans, bank-specific) would map to the wrong field for
            # non-bank companies. Dispatcher selects the correct dict from detected sheet type.
            headers_dict = HEADERS_BY_TYPE[effective_type]
            debug_mode = effective_type in (
                "insurance_balance_sheet", "insurance_income_statement",
                "bank_balance_sheet", "bank_income_statement",
                "securities_balance_sheet", "securities_income_statement",
            )
            parsed_data = _parse_openpyxl_sheet(ws, headers_dict, debug=debug_mode)

            if parsed_data:
                # Merge into existing data (don't overwrite) so a supplementary sheet like
                # Тодруулга (notes) adds new fields without clobbering the primary balance sheet.
                existing = result.get(effective_type, {})
                existing.update({k: v for k, v in parsed_data.items() if k not in existing})
                result[effective_type] = existing
                if effective_type not in result["metadata"]["sheets_parsed"]:
                    result["metadata"]["sheets_parsed"].append(effective_type)

            # Extract company name from sheet content (always prefer over filename)
            extracted = _extract_company_from_openpyxl_sheet(ws)
            if extracted:
                result["metadata"]["company"] = extracted
    finally:
        wb.close()

    if not result["metadata"]["sheets_parsed"]:
        raise ValueError(
            "No recognizable financial statement sheets found. "
            "Expected Mongolian-language sheets (e.g., Балансын тайлан, "
            "Орлогын тайлан, Мөнгөн гүйлгээний тайлан)."
        )

    return result


def _parse_xls(file_bytes: bytes, filename: str, sector: str = "") -> dict:
    """Parse a .xls file using xlrd."""
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
    except Exception as e:
        raise ValueError(f"Cannot open Excel file: {e}")

    company, year = _extract_metadata_from_filename(filename)

    # Detect reporting unit from first sheet before building result dict
    unit_multiplier = 1_000
    if wb.sheet_names():
        unit_multiplier = _detect_reporting_unit_xlrd(wb.sheet_by_index(0))

    result = _make_result_dict(filename, company, year, unit_multiplier)

    # Pre-scan sheets for company name so sector lookup works even when the
    # filename is an opaque MSE code (e.g. "56320254report.xls").
    if not sector:
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            extracted = _extract_company_from_xlrd_sheet(ws)
            if extracted:
                company = extracted
                result["metadata"]["company"] = extracted
                break

    resolved_sector_key = _lookup_sector_key(company, filename) if not sector else sector
    sector_overrides = _SECTOR_SHEET_OVERRIDES.get(resolved_sector_key, {})

    for sheet_name in wb.sheet_names():
        sheet_type = _detect_sheet_type(sheet_name)
        if not sheet_type:
            continue

        effective_type = sector_overrides.get(sheet_type, sheet_type)

        ws = wb.sheet_by_name(sheet_name)
        headers_dict = HEADERS_BY_TYPE[effective_type]
        debug_mode = effective_type in (
            "insurance_balance_sheet", "insurance_income_statement",
            "bank_balance_sheet", "bank_income_statement",
            "securities_balance_sheet", "securities_income_statement",
        )
        parsed_data = _parse_xlrd_sheet(ws, headers_dict, debug=debug_mode)

        if parsed_data:
            # Merge: don't overwrite existing keys so Тодруулга adds new fields only.
            existing = result.get(effective_type, {})
            existing.update({k: v for k, v in parsed_data.items() if k not in existing})
            result[effective_type] = existing
            if effective_type not in result["metadata"]["sheets_parsed"]:
                result["metadata"]["sheets_parsed"].append(effective_type)

        # Extract company name from sheet content (always prefer over filename)
        extracted = _extract_company_from_xlrd_sheet(ws)
        if extracted:
            result["metadata"]["company"] = extracted

    if not result["metadata"]["sheets_parsed"]:
        raise ValueError(
            "No recognizable financial statement sheets found. "
            "Expected Mongolian-language sheets (e.g., Балансын тайлан, "
            "Орлогын тайлан, Мөнгөн гүйлгээний тайлан)."
        )

    return result


def _make_result_dict(filename: str, company: str, year: str, unit_multiplier: int = 1_000) -> dict:
    """Create the base result dictionary."""
    return {
        "metadata": {
            "filename": filename,
            "company": company,
            "year": year,
            "parsed_at": datetime.now().isoformat(),
            "sheets_parsed": [],
            "reporting_unit_multiplier": unit_multiplier,
        },
    }


def _extract_metadata_from_filename(filename: str) -> tuple[str, str]:
    """Extract company name and year from filename.

    Supports patterns like:
        APU_2023.xlsx -> ("APU", "2023")
        APU 2023.xlsx -> ("APU", "2023")
        354_finance_report_xls_698c1feac88ac.xls -> ("unknown", "unknown")
    """
    stem = filename.rsplit(".", 1)[0] if "." in filename else filename

    year_match = re.search(r"(20\d{2})", stem)
    year = year_match.group(1) if year_match else "unknown"

    if year_match:
        company = stem[: year_match.start()].strip().rstrip("_- ")
    else:
        company = stem.strip()

    # If company looks like a hash or random string, treat as unknown
    if company and re.match(r"^[\da-f_]+$", company):
        company = "unknown"

    return company if company else "unknown", year


def _detect_reporting_unit_multiplier_from_text(text: str) -> int | None:
    """Return the numeric multiplier implied by a unit declaration cell.

    Mongolian MSE files declare the reporting unit in the first few rows, e.g.:
      /Мянган төгрөг/   → 1 000 (thousands)
      /мян.төгрөгөөр/   → 1 000
      /Сая төгрөг/      → 1 000 000 (millions)
      / төгрөг/         → 1 (raw MNT)

    Returns the multiplier (1, 1000, or 1_000_000) or None if text doesn't
    look like a unit declaration.
    """
    if not text:
        return None
    t = str(text).strip().lower()
    if "мян" in t:
        return 1_000
    if "сая" in t:
        return 1_000_000
    # raw MNT: contains "төгрөг" but no мян/сая qualifier
    if "төгрөг" in t:
        return 1
    return None


def _detect_reporting_unit_openpyxl(ws) -> int:
    """Scan first 8 rows of an openpyxl sheet for the unit declaration."""
    for row in ws.iter_rows(min_row=1, max_row=8, max_col=5):
        for cell in row:
            if cell.value:
                mult = _detect_reporting_unit_multiplier_from_text(str(cell.value))
                if mult is not None:
                    return mult
    return 1_000  # default: most MSE companies report in thousands


def _detect_reporting_unit_xlrd(ws) -> int:
    """Scan first 8 rows of an xlrd sheet for the unit declaration."""
    for row_idx in range(min(8, ws.nrows)):
        for col_idx in range(min(5, ws.ncols)):
            val = ws.cell_value(row_idx, col_idx)
            if val:
                mult = _detect_reporting_unit_multiplier_from_text(str(val))
                if mult is not None:
                    return mult
    return 1_000  # default: most MSE companies report in thousands


def _extract_company_name_from_text(text: str) -> str | None:
    """Extract company name from cell text like 'Байгууллагын нэр: Сүү'."""
    if not text:
        return None
    text = str(text).strip()
    match = re.search(r"байгууллагын нэр\s*:\s*(.+)", text, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return None


def _extract_company_from_openpyxl_sheet(ws) -> str | None:
    """Scan first few rows of an openpyxl sheet for company name."""
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=5):
        for cell in row:
            if cell.value:
                name = _extract_company_name_from_text(cell.value)
                if name:
                    return name
    return None


def _extract_company_from_xlrd_sheet(ws) -> str | None:
    """Scan first few rows of an xlrd sheet for company name."""
    for row_idx in range(min(5, ws.nrows)):
        for col_idx in range(min(5, ws.ncols)):
            val = ws.cell_value(row_idx, col_idx)
            if val:
                name = _extract_company_name_from_text(val)
                if name:
                    return name
    return None


def _detect_sheet_type(sheet_name: str) -> str | None:
    """Detect the financial statement type from sheet name.

    # Sheet type is detected from tab name patterns before header mapping:
    # determines which of the 7 dictionaries (standard BS/IS/CF, bank BS/IS,
    # insurance BS/IS) to apply. Companies filing under different regulatory
    # frameworks use entirely different Mongolian account terminology.
    """
    normalized = normalize_header(sheet_name)
    # Check longer patterns first for specificity
    for pattern, sheet_type in sorted(
        SHEET_TYPE_PATTERNS.items(), key=lambda x: len(x[0]), reverse=True
    ):
        if pattern in normalized:
            return sheet_type
    return None


def _parse_openpyxl_sheet(ws, headers_dict: dict[str, str], debug: bool = False) -> dict:
    """Parse an openpyxl worksheet (.xlsx).

    Scans column C (index 2) for Mongolian headers, reads values
    from columns D (index 3 = prev year) and E (index 4 = current year).
    """
    parsed = {}

    for row in ws.iter_rows(min_col=1, max_col=6):
        # Get header text from column C (index 2)
        if len(row) <= HEADER_COL:
            continue
        header_cell = row[HEADER_COL]
        if header_cell.value is None:
            continue

        english_key = match_header(header_cell.value, headers_dict)
        if english_key is None:
            if debug:
                prev_val = row[PREV_COL].value if len(row) > PREV_COL else ""
                curr_val = row[CURR_COL].value if len(row) > CURR_COL else ""
                logging.warning("UNMATCHED HEADER: %r | prev=%r | curr=%r", header_cell.value, prev_val, curr_val)
            continue

        # Read previous year (col D) and current year (col E)
        prev_val = _safe_cell_value(row, PREV_COL)
        curr_val = _safe_cell_value(row, CURR_COL)

        # First-match-wins: the total/summary row always appears before sub-items in
        # MSE filing format. Sub-items contain the same Mongolian phrase as the parent
        # (e.g. "зээлийн хүүгийн орлого" matches both the total and each loan category),
        # so we keep the first (total) value and ignore subsequent sub-item matches.
        # Exception: a section-header row that lands as 0.0 (placeholder) may appear
        # before the real total row (e.g. "Эздийн өмч" before "Эздийн өмчийн дүн").
        # Allow a non-zero value to override a previously stored 0.0 so the real total wins.
        if curr_val is not None:
            if english_key not in parsed or (parsed[english_key] == 0.0 and curr_val != 0.0):
                parsed[english_key] = curr_val
        prev_key = f"{english_key}_prev"
        if prev_val is not None:
            if prev_key not in parsed or (parsed.get(prev_key) == 0.0 and prev_val != 0.0):
                parsed[prev_key] = prev_val

    return parsed


def _parse_xlrd_sheet(ws, headers_dict: dict[str, str], debug: bool = False) -> dict:
    """Parse an xlrd worksheet (.xls).

    Scans column C (index 2) for Mongolian headers, reads values
    from columns D (index 3 = prev year) and E (index 4 = current year).
    """
    parsed = {}

    for row_idx in range(ws.nrows):
        # Get header text from column C (index 2)
        if ws.ncols <= HEADER_COL:
            continue
        header_val = ws.cell_value(row_idx, HEADER_COL)
        if not header_val:
            continue

        english_key = match_header(header_val, headers_dict)
        if english_key is None:
            if debug:
                prev_val = ws.cell_value(row_idx, PREV_COL) if ws.ncols > PREV_COL else ""
                curr_val = ws.cell_value(row_idx, CURR_COL) if ws.ncols > CURR_COL else ""
                logging.warning("UNMATCHED HEADER: %r | prev=%r | curr=%r", header_val, prev_val, curr_val)
            continue

        # Read previous year (col D) and current year (col E)
        prev_val = _safe_xlrd_value(ws, row_idx, PREV_COL)
        curr_val = _safe_xlrd_value(ws, row_idx, CURR_COL)

        # First-match-wins: keeps the total/summary row and ignores subsequent sub-items
        # that share the same Mongolian phrase (see openpyxl variant for full explanation).
        # Exception: allow non-zero to override a 0.0 placeholder from a section header row.
        if curr_val is not None:
            if english_key not in parsed or (parsed[english_key] == 0.0 and curr_val != 0.0):
                parsed[english_key] = curr_val
        prev_key = f"{english_key}_prev"
        if prev_val is not None:
            if prev_key not in parsed or (parsed.get(prev_key) == 0.0 and prev_val != 0.0):
                parsed[prev_key] = prev_val

    return parsed


def _safe_cell_value(row, col_idx: int) -> float | None:
    """Safely get a numeric value from an openpyxl row at col_idx."""
    if col_idx >= len(row):
        return None
    val = row[col_idx].value
    if val is None:
        return None
    if _is_numeric(val):
        return _to_number(val)
    return None


def _safe_xlrd_value(ws, row_idx: int, col_idx: int) -> float | None:
    """Safely get a numeric value from an xlrd sheet."""
    if col_idx >= ws.ncols:
        return None
    val = ws.cell_value(row_idx, col_idx)
    if val is None or val == "":
        return None
    if _is_numeric(val):
        return _to_number(val)
    return None


def _is_numeric(value) -> bool:
    """Check if a value is numeric (int, float, or numeric string)."""
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace(" ", "").strip()
        try:
            float(cleaned)
            return True
        except ValueError:
            return False
    return False


def _to_number(value) -> float:
    """Convert a value to a number."""
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").replace(" ", "").strip()
    return float(cleaned)
